import PyQt5
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QPushButton, QWidget, QLineEdit
from PyQt5.QtWidgets import QTableWidgetItem
from PyQt5.QtWidgets import  QInputDialog, QFileDialog
from PyQt5.QtGui import QIcon
from PyQt5.QtGui import QKeyEvent
from PyQt5.QtCore import Qt
from importhelper import Ui_MainWindow
from PyQt5.QtGui import QPalette, QColor  # для цветов
from PyQt5 import QtCore

import sys
import os
import locale
from pycbrf.toolbox import ExchangeRates
from datetime import datetime
from StringStorage import storeString
import openpyxl                         #модуль работы с эксель
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook





class mywindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(mywindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.memory = "0"
        self.now = datetime.now()                               # получаем дату сегодня
        self.ui.DateToday.setText(str(self.now.date()))         # пишем дату в приложоение
        self.goodsNumber = 0                                    # делаем переменную на кол-во товаров
        self.sumTax = 0                                         # делаем переменную на пошлину
        self.sumVAT = 0                                         # делаем переменную на  НДС
        self.invoiceSum = 0                                     # делаем переменную на сумму инвойса
        self.taxValue = 0                                       # делаем переменную на выовод пошлины  до преобразование в разряды
        self.VATvalue = 0                                       # делаем переменную на вывод НДС до преобразование в разряды
        self.VATvalue_out = 0                                   # делаем переменную на выовод пошлины  после преобразование в разряды
        self.taxValue_out = 0                                   # делаем переменную на вывод НДС после преобразование в разряды
        self.taxValue_beforeShow = 0                            # делаем переменную на расчет пошлины  и округления до 2 знака
        self.VATvalue_beforeShow = 0                            # делаем переменную на расчет НДС  и округления до 2 знака
        self.save1 = ""                                         # делаем переменную на полученной значение из строки СУММА ИНВОЙСА
        self.save2 = ""                                         # делаем переменную на полученной значение из строки СУММА ТРАНСПОРТА
        self.taxInfo = ""                                       # делаем переменную на полученной значение из строки ПОШЛИНА
        self.save1RUB = 0                                       # делаем переменную на полученной значение из строки СУММА ИНВОЙСА в РУБЛЯХ
        self.save2RUB = 0                                       # делаем переменную на полученной значение из строки СУММА ТРАНСПОРТА в РУБЛЯХ
        self.invoiceSum_out = 0                                 # перменная для вывода суммы инвойса в память с разрядами 
        self.sumTax_out = 0                                     # перменная для вывода суммы пошлины в память с разрядами 
        self.sumVAT_out = 0                                     # перменная для вывода суммы НДС в память с разрядами    
        self.combiTax = ""                                      # перменная для ввода сохранения комбинированной пошлины   
        self.combiTaxRUB = ""                                   # перменная для ввода сохранения комбинированной пошлины в рублях                                 
        self.nettoWeight = ""                                   # перменная для ввода сохранения веса нетто   
        self.totalTax = 0                                       # переменная для сохранения полной суммы налогов по ДТ
        self.totalTaxValue = 0
        self.totalTaxValue_out = 0                              # перменная для вывода суммы налогов в память с разрядами    
        self.totalStringOut = ""                                 #переменная куда мы сохраняем расшифровку по товарам
        self.invoiceCurrency = ""
        self.paymentsCurrency = ""
        self.sumTransport = 0
        self.tranportSum_out = 0
        self.transportCurrency =""
        self.file_name = ""
        self.uchet = 0                                      # фЛАГ чтобы не менять валюту рассчета платежей больше 1 раза на каждый товар
        self.sbor = 0                                       # переменная для таможенных сборов
        self.sborCurrency = 0
        self.sborOut = 0
        self.save1RUBCommulative = 0
        self.totalInvoicePayment = 0                        # переменная для рассчета суммы платежей по инвойсу 
        self.totalInvoicePaymentOut = 0                     # переменная для вывода суммы платежей по инвойсу в формате РФ



        self.ui.tableWidget.setColumnCount(9)
        self.ui.tableWidget.setHorizontalHeaderLabels(('ИНВОЙС', "Вал.", "%", "Пошлина", "Вал.", "НДС", "Вал.", "Тр-т", "Вал."))   #задаем на таблички справа назване колонок
        self.ui.tableWidget.setColumnWidth(0, 60)                                       # ниже задаем размеры каждой колонки
        self.ui.tableWidget.setColumnWidth(1, 20)
        self.ui.tableWidget.setColumnWidth(2, 20)
        self.ui.tableWidget.setColumnWidth(3, 70)
        self.ui.tableWidget.setColumnWidth(4, 20)
        self.ui.tableWidget.setColumnWidth(5, 70)
        self.ui.tableWidget.setColumnWidth(6, 20)
        self.ui.tableWidget.setColumnWidth(7, 60)
        self.ui.tableWidget.setColumnWidth(8, 20)    
              
        self.ui.tableWidget.setRowCount(100)                            # задаем количество строк
        
        locale.setlocale(locale.LC_ALL, "ru")                   # устанавлием стандарты РФ 
        
        self.data = []                              # создаем массив куда будем хранить товары в массиве.

        i = 0                                   
        for i in range(100):                        # Добавляем 100 строк в товары в массиве.
            self.data.append([0]*10)         

        self.data[0] = ['№','ИНВОЙС', "Вал.", "%", "Пошлина", "Вал.", "НДС", "Вал.", "Тр-т", "Вал."]   #  первая строка в массиве хранения 
                                                                                            # товаров будем хранить наименование столбиков.
        
        

        self.rates = ExchangeRates(self.now.date(), locale_en=True)     # используем библиотеку получаем курсы валют
        self.rates.date_requested # 2016-06-26 00:00:00
        self.rates.date_received # 2016-06-25 00:00:00
        self.rates.dates_match # False


        self.valueUSD = self.rates['USD'].value                     # присваиваем полученный курс доллара
        self.valueEUR = self.rates['EUR'].value                     # присваиваем полученный курс Евро
        self.valueCNY = self.rates['CNY'].value                     # присваиваем полученный курс Юань
        #self.valueUSD = 75
        #self.valueEUR = 85
        #self.valueCNY = 10

#print(valueUSD, valueEUR, valueCNY)
        

# так задаем курсы
        self.ui.USDToday.setText(str(self.valueUSD))                     # записываем полученный курс доллара
        self.ui.EURToday.setText(str(self.valueEUR))                     # записываем полученный курс Евро
        self.ui.CNYToday.setText(str(self.valueCNY))                     # записываем полученный курс Юань
#print(valueUSD, valueEUR, valueCNY)

        #окно для ввода №1 инвойс
        self.textbox = QLineEdit(self)                              # создаем кьюлайнэдит
        self.textbox.move(110, 230)                                 # создаем начальные координаты х, у
        self.textbox.resize(160,40)                                 # создаем конечные  координаты х, у
    #    self.textbox.setText("0")

        #окно для ввода №3 пошлина
        self.textbox3 = QLineEdit(self)                             # создаем кьюлайнэдит
        self.textbox3.move(400, 230)                                # создаем начальные координаты х, у
        self.textbox3.resize(140,40)                                # создаем конечные  координаты х, у
        self.textbox3.setText("0")                                  # сразу пишем туда ноль чтобы все работало

        #окно для ввода №2 трансопрт
        self.textbox2 = QLineEdit(self)                             # создаем кьюлайнэдит
        self.textbox2.move(110, 300)                                # создаем начальные координаты х, у
        self.textbox2.resize(160,40)                                # создаем конечные  координаты х, у
        self.textbox2.setText("0")                                  # сразу пишем туда ноль чтобы все работало



        # окно для ввода комбинированной пошлины
        self.textbox4 = QLineEdit(self)                             # создаем кьюлайнэдит
        self.textbox4.move(400, 300)                                # создаем начальные координаты х, у
        self.textbox4.resize(60,40)                                 # создаем конечные  координаты х, у
        self.textbox4.setText("0")                                  # сразу пишем туда ноль чтобы все работало

        # окно для ввода веса нетто товара
        self.textbox5 = QLineEdit(self)                             # создаем кьюлайнэдит
        self.textbox5.move(400, 370)                                # создаем начальные координаты х, у
        self.textbox5.resize(140,40)                                # создаем конечные  координаты х, у
        self.textbox5.setText("0")                                  # сразу пишем туда ноль чтобы все работало



        # тут описываем кнопки 
        self.ui.pushButtonEqual.clicked.connect(self.pushButtonEqual)               # подключение клик-сигнал к слоту btnClicked  включаем мышку через коннект и запускаем слот на равно   
        self.ui.pushButtonMemory.clicked.connect(self.pushButtonMemory)             # подключение клик-сигнал к слоту btnClicked  включаем мышку через коннект и запускаем слот на добавление в память 
        self.ui.pushButtonMemoryClean.clicked.connect(self.pushButtonMemoryClean)   # подключение клик-сигнал к слоту btnClicked  включаем мышку через коннект и запускаем слот очистку памяти
        self.ui.pushButtonExcelExport.clicked.connect(self.pushButtonExcelExport) 
        self.ui.pushButtonOpenFile.clicked.connect(self.pushButtonOpenFile) 

        # включаем клавиши и сравниваем аски. если верно то запускаем ту же функцию что и на мышку
    def keyPressEvent(self, e):     
#        print(e.key())
        if e.key() == 16777220 or e.key() == 16777221:              # если нажимают большой или малый ентер, то запускаем слот на равно 
            self.pushButtonEqual()
        if e.key() == 16777265:                                     # если нажимают Ф2, то запускаем слот на добавление в память
            self.pushButtonMemory()                 
        
        # описываю слот на нажатие рассчета
    def pushButtonEqual(self):
        self.save1 = float(storeString(self.textbox.text()))    # получаем цену товара
        self.save2 = float(storeString(self.textbox2.text()))   #  получаем цену транспорта
        self.taxInfo = float(storeString(self.textbox3.text())) # получаем пошлину
        self.combiTax = float(storeString(self.textbox4.text()))  # получаем комбинированную пошлину
        self.nettoWeight = float(storeString(self.textbox5.text()))  # получаем вес нетто

# описываем комбобокс один на 
        if self.ui.comboBox.currentText() == "USD":                         # если доллары в знач. сумму товара
            self.save1RUB = float(self.save1) * float(self.valueUSD)        # то переводим в рубли сумму инвойс
            self.ui.MemoryInvoiceSum_2.setText("USD")                       # в память запишем валюту инвойса доллар
            self.invoiceCurrency = "USD"
            if self.uchet == 0:
                self.ui.comboBox_TAXtoPay.setCurrentIndex(0)
                self.ui.comboBox_VATtoPAY.setCurrentIndex(0)
                self.ui.comboBox_VATtoPAY_2.setCurrentIndex(0)

                self.uchet +=1


        if self.ui.comboBox.currentText() == "EUR":                         # сумма в Евро
            self.save1RUB = float(self.save1) * float(self.valueEUR)        # 
            self.ui.MemoryInvoiceSum_2.setText("EUR")
            self.invoiceCurrency = "EUR"
            if self.uchet == 0:
                self.ui.comboBox_TAXtoPay.setCurrentIndex(2)
                self.ui.comboBox_VATtoPAY.setCurrentIndex(2)
                self.ui.comboBox_VATtoPAY_2.setCurrentIndex(2)

                self.uchet +=1

        if self.ui.comboBox.currentText() == "CNY":                         # сумма в Юани
            self.save1RUB = float(self.save1) * float(self.valueCNY)
            self.ui.MemoryInvoiceSum_2.setText("CNY")
            self.invoiceCurrency = "CNY"
            if self.uchet == 0:
                self.ui.comboBox_TAXtoPay.setCurrentIndex(3)
                self.ui.comboBox_VATtoPAY.setCurrentIndex(3)
                self.ui.comboBox_VATtoPAY_2.setCurrentIndex(3)
                self.uchet +=1


# описываем комбобокс 2 на
        if self.ui.comboBox_2.currentText() == "RUB":                       # переводим стоимость транспорта в рубли
            self.save2RUB = float(self.save2)
            self.ui.MemoryTaxSum_4.setText("RUB")                               # и присваиваем соотвествующей переменной
            self.transportCurrency = "RUB"


        if self.ui.comboBox_2.currentText() == "USD":                       # переводим стоимость транспорта в рубли из Доллвроы
            self.save2RUB = float(self.save2) * float(self.valueUSD)
            self.ui.MemoryTaxSum_4.setText("USD")
            self.transportCurrency = "USD"

        if self.ui.comboBox_2.currentText() == "EUR":                       # переводим стоимость транспорта в рубли из Евро
            self.save2RUB = float(self.save2) * float(self.valueEUR)
            self.ui.MemoryTaxSum_4.setText("EUR")
            self.transportCurrency = "EUR"

        if self.ui.comboBox_2.currentText() == "CNY":                       # переводим стоимость транспорта в рубли из Юаней
            self.save2RUB = float(self.save2) * float(self.valueCNY)
            self.ui.MemoryTaxSum_4.setText("CNY")
            self.transportCurrency = "CNY"

# описываем комбобокс 3 на
        if self.ui.comboBox_3.currentText() == "EUR/kg":                    # переводим стоимость комбинированной пошлины в если комбобокс евро за кг 
            self.combiTaxRUB = float(self.combiTax) * float(self.valueEUR)  # если евро в рубли


        if self.ui.comboBox_3.currentText() == "USD/kg":                    # переводим стоимость комбинированной пошлины в если комбобокс Долл США за кг 
            self.combiTaxRUB = float(self.combiTax) * float(self.valueUSD)  # если США в рубли

        if self.ui.comboBox_3.currentText() == "CNY/kg":                    # переводим стоимость комбинированной пошлины в если комбобокс Долл Юане за кг
            self.combiTaxRUB = float(self.combiTax) * float(self.valueCNY)  # если Юаней в рубли
 


        self.taxValue_beforeShow = round((self.save1RUB+self.save2RUB)*(self.taxInfo/100) + self.nettoWeight * self.combiTaxRUB, 2)     # считаем пошлину в рублях
        self.VATvalue_beforeShow = round((self.taxValue_beforeShow + self.save1RUB + self.save2RUB)*0.2, 2) # считаем НДС в рублях
        self.totalTax = self.taxValue_beforeShow + self.VATvalue_beforeShow                                   # считаем сумму платежей + сборов


# описываем комбобокс для вывода валюты пошлины

        if self.ui.comboBox_TAXtoPay.currentText() == "RUB":                    # если комбобокс руб 
            self.taxValue = float(self.taxValue_beforeShow)                     # тогда платежи  не умножаем на курс
            self.ui.MemoryTaxSum_2.setText("RUB")                               # и пишем в валюту графы памяти 
            self.paymentsCurrency = "RUB"
           

        if self.ui.comboBox_TAXtoPay.currentText() == "USD":                         # если комбобокс Долл США 
            self.taxValue = float(self.taxValue_beforeShow) / float(self.valueUSD)   # тогда платежи  умножаем на курс Доллара
            self.ui.MemoryTaxSum_2.setText("USD")                                    # и пишем в валюту графы памяти 
            self.paymentsCurrency = "USD"
 

        if self.ui.comboBox_TAXtoPay.currentText() == "EUR":                         # если комбобокс Евро 
            self.taxValue = float(self.taxValue_beforeShow) / float(self.valueEUR)
            self.ui.MemoryTaxSum_2.setText("EUR")
            self.paymentsCurrency = "EUR" 


        if self.ui.comboBox_TAXtoPay.currentText() == "CNY":                         # если комбобокс Юани 
            self.taxValue = float(self.taxValue_beforeShow) / float(self.valueCNY)
            self.ui.MemoryTaxSum_2.setText("CNY")
            self.paymentsCurrency = "CNY" 

# описываем комбобокс для вывода валюты НДС

        if self.ui.comboBox_VATtoPAY.currentText() == "RUB":
            self.VATvalue = float(self.VATvalue_beforeShow)
            self.ui.MemoryTaxSum_3.setText("RUB")
            self.ui.MemoryTaxSum_6.setText("RUB")
            self.ui.MemoryTaxSum_5.setText("RUB") 


        if self.ui.comboBox_VATtoPAY.currentText() == "USD":
            self.VATvalue = float(self.VATvalue_beforeShow) / float(self.valueUSD)
            self.ui.MemoryTaxSum_3.setText("USD") 
            self.ui.MemoryTaxSum_6.setText("USD") 
            self.ui.MemoryTaxSum_5.setText("USD") 

        if self.ui.comboBox_VATtoPAY.currentText() == "EUR":
            self.VATvalue = float(self.VATvalue_beforeShow) / float(self.valueEUR)
            self.ui.MemoryTaxSum_3.setText("EUR") 
            self.ui.MemoryTaxSum_6.setText("EUR") 
            self.ui.MemoryTaxSum_5.setText("EUR") 

        if self.ui.comboBox_VATtoPAY.currentText() == "CNY":
            self.VATvalue = float(self.VATvalue_beforeShow) / float(self.valueCNY)
            self.ui.MemoryTaxSum_3.setText("CNY") 
            self.ui.MemoryTaxSum_6.setText("CNY") 
            self.ui.MemoryTaxSum_5.setText("CNY") 

# описываем комбобокс для вывода валюты всего

        if self.ui.comboBox_VATtoPAY_2.currentText() == "RUB":
            self.totalTaxValue = float(self.totalTax)
 

        if self.ui.comboBox_VATtoPAY_2.currentText() == "USD":
            self.totalTaxValue = float(self.totalTax) / float(self.valueUSD)
 

        if self.ui.comboBox_VATtoPAY_2.currentText() == "EUR":
            self.totalTaxValue = float(self.totalTax) / float(self.valueEUR)


        if self.ui.comboBox_VATtoPAY_2.currentText() == "CNY":
            self.totalTaxValue = float(self.totalTax) / float(self.valueCNY)



        self.taxValue_out = locale.format_string('%.2f', self.taxValue, True)               # преобразуем пошлину в формат с разрядами
        self.VATvalue_out = locale.format_string('%.2f', self.VATvalue, True)               # преобразуем НДС в формат с разрядами
        self.totalTaxValue_out = locale.format_string('%.2f', self.totalTaxValue, True)     # преобразуем пошлину+НДС в формат с разрядами
        
        self.ui.TAXtoPay.setText(str(self.taxValue_out))                        #  пишем в строчки пошлину
        self.ui.VATtoPAY.setText(str(self.VATvalue_out))                        # НДС
        self.ui.totalTaxes.setText(str(self.totalTaxValue_out))                 # сумму платежей     
                                  # пишем рубли в сбор
         #

        if self.save2 == 0:
            self.textbox2.setStyleSheet("background-color: rgb(255, 240, 240);")
        else:
            self.textbox2.setStyleSheet("background-color: rgb(255, 255, 255);")
   
        if self.taxInfo == 0:
            self.textbox3.setStyleSheet("background-color: rgb(255, 240, 240);")
        else:
            self.textbox3.setStyleSheet("background-color: rgb(255, 255, 255);")


        return self.taxValue, self.VATvalue, self.save1, self.invoiceCurrency, self.save2, self.transportCurrency     # возвращаем пошлину, НДС и сумму инвойса и его валюту

    def pushButtonMemory(self):                          # Добавление в память
        self.pushButtonEqual()
        if self.save1 != 0:
    
  
                data = [0]*9                       # если инвойс не равен нулю то делаем строчку массива 
                self.goodsNumber += 1                       # счетчик товаров
                self.invoiceSum += float(self.save1)        # добавляем к накопленной сумме инвойса еще одну
                self.sumTax += float(self.taxValue)         # добавляем к накопленной сумме пошлины еще одну
                self.sumVAT += float(self.VATvalue)         # добавляем к напопленной сумме ндс еще одну
                self.sumTransport+= float(self.save2)
                self.save1RUBCommulative += self.save1RUB
                if self.save1RUBCommulative < 200000:
                    self.sbor = 750
                if self.save1RUBCommulative > 200000 and self.save1RUBCommulative <= 450000:
                    self.sbor = 1550
                if self.save1RUBCommulative > 450000 and self.save1RUBCommulative <= 1200000:
                    self.sbor = 3100
                if self.save1RUBCommulative > 1200000 and self.save1RUBCommulative <= 2700000:
                    self.sbor = 8530
                if self.save1RUBCommulative > 2700000 and self.save1RUBCommulative <= 4200000:
                    self.sbor = 12000
                if self.save1RUBCommulative > 4200000 and self.save1RUBCommulative <= 5500000:
                    self.sbor = 15500
                if self.save1RUBCommulative > 5500000 and self.save1RUBCommulative <= 7000000:
                    self.sbor = 20000
                if self.save1RUBCommulative > 7000000 and self.save1RUBCommulative <= 8000000:
                    self.sbor = 23000
                if self.save1RUBCommulative > 8000000 and self.save1RUBCommulative <= 9000000:
                    self.sbor = 25000
                if self.save1RUBCommulative > 9000000 and self.save1RUBCommulative <= 10000000:
                    self.sbor = 27000
                if self.save1RUBCommulative > 10000000:
                    self.sbor = 30000
                
                if self.ui.comboBox_VATtoPAY.currentText() == "RUB":

                    self.sborCurrency = self.sbor

                if self.ui.comboBox_VATtoPAY.currentText() == "USD":
                    self.sborCurrency = float(self.sbor) / float(self.valueUSD) 
                
                if self.ui.comboBox_VATtoPAY.currentText() == "EUR":
                    self.sborCurrency = float(self.sbor) / float(self.valueEUR) 
                
                if self.ui.comboBox_VATtoPAY.currentText() == "CNY":
                    self.sborCurrency = float(self.sbor) / float(self.valueCNY) 

                self.totalInvoicePayment = (self.sumTax + self.sumVAT + self.sborCurrency)   # рассчитываем сумму всего платежей по инвойсу

                # НУЛИМ СТАРОЕ------------------------------------------------------------- 

                self.invoiceSum_out = locale.format_string('%.2f', self.invoiceSum, True) # преобразуем инвойс в памяти в формат с разрядами
                self.sumTax_out = locale.format_string('%.2f', self.sumTax, True)          # преобразуем пошлину в памяти в формат с разрядами
                self.sumVAT_out = locale.format_string('%.2f', self.sumVAT, True)           # преобразуем НДС в памяти в формат с разрядами
                self.tranportSum_out = locale.format_string('%.2f', self.sumTransport, True)
                self.sborOut = locale.format_string('%.2f', self.sborCurrency, True)    
                self.totalInvoicePaymentOut = locale.format_string('%.2f', self.totalInvoicePayment, True)    
                           

                
                self.ui.MemoryNumberofGoods.setText(str(self.goodsNumber))              # пишем в память номер товара
                self.ui.MemoryInvoiceSum.setText(str(self.invoiceSum_out))              # пишем в память накопленную сумму инвойс
                self.ui.MemoryTaxSum.setText(str(self.sumTax_out))                      # пишем в память накопленную сумму пошлины
                self.ui.MemoryVatSum.setText(str(self.sumVAT_out))                      # пишем в память накопленную сумму НДС
                self.ui.MemoryTransportCost.setText(str(self.tranportSum_out))
                self.ui.MemorySbor.setText(str(self.sborOut))
                self.ui.MemoryVatSum_2.setText(str(self.totalInvoicePaymentOut))        # пишем в памяти сумму платежей по инвойсу

                self.taxValue = round(self.taxValue, 2)                             # округляем пошлину для записи в лог
                self.VATvalue = round(self.VATvalue, 2)                             # округляем НДС для записи в лог




                data = [str(self.goodsNumber), str(self.save1), str(self.invoiceCurrency), str(self.taxInfo), str(self.taxValue), 
                str(self.paymentsCurrency), str(self.VATvalue), str(self.paymentsCurrency), str(self.save2), str(self.transportCurrency)]   # заполняем массив данными по строчке
                
                i = 0
                y = self.goodsNumber-1                         # т.к. вначале мы уже счелкнули товар, то приходится минус 1 делать
                while i<9:                                      # заполняем поячеечно строчку лога
                    it1 = QTableWidgetItem(data[i+1])           # получаем адрес в какой объект пишем
                    self.ui.tableWidget.setItem(y, i, it1)      # и пишем адрес и данные
                    i += 1

                self.data[self.goodsNumber] = data              # также данные по строке пишем в лог матрицу





                #нулим то что было

                self.taxValue = 0                           # пошлнину 
                self.VATvalue = 0                           # НДС
                self.save1 = 0                              # инвойс
                self.combiTaxRUB = 0                        # комбинированную пошлину
                self.nettoWeight = 0                        # вес нетто
                self.totalTax = 0                           # сумму платежей
            #    self.uchet = 0


                self.textbox.setText("")                                                # нулим строку с инвойсом 
                self.textbox2.setText("0")                                              # нулим строку с транспортом
                self.textbox3.setText("0")                                              # нулим строку с пошлиной
                self.textbox4.setText("0")                                              # нулим строку с комбинированной пошлиной
                self.textbox5.setText("0")                                              # нулим строку с весом нетто

                self.ui.TAXtoPay.setText(str(self.taxValue))                            # устанавлиеваем ноль в строку платежей пошлины 
                self.ui.VATtoPAY.setText(str(self.VATvalue))                            # устанавливаем ноль в строку платежей ндс
                self.ui.totalTaxes.setText(str(self.totalTax))                          # устанавлиеваем ноль в строку платежей итого



        return self.goodsNumber, self.invoiceSum, self.sumTax, self.sumVAT, self.taxValue, self.VATvalue, self.save1, self.data, self.uchet   # возвращаем нули в основную функцию

    def pushButtonMemoryClean(self):                                        # функция чистим память
        self.goodsNumber = 0                                                # функция нулим номар товара
        self.invoiceSum = 0                                                 # нулим значение инвойса
        self.sumTax = 0                                                     # нулим пошлину в памяти
        self.sumVAT = 0                                                     # нулим НДС в памяти
        self.sumTransport = 0
        self.tranportSum_out = 0
        self.uchet = 0
        self.sbor = 0                                                         # нулим сбор в памяти
        self.save1RUBCommulative = 0
        self.totalInvoicePayment = 0
        self.totalInvoicePaymentOut = 0


        self.ui.MemoryNumberofGoods.setText(str(self.goodsNumber))          # пишем в строчку товаров номер товара - НОЛЬ
        self.ui.MemoryInvoiceSum.setText(str(self.invoiceSum))              # пишем в строчку инвойса - НОЛЬ
        self.ui.MemoryTaxSum.setText(str(self.sumTax))                      # пишем в строчку пошлины НОЛЬ
        self.ui.MemoryVatSum.setText(str(self.sumVAT))                      # пишем в строчку НДС - НОЛЬ
        self.ui.MemorySbor.setText(str(self.sbor))                          # пишем в строчку сбора - НОЛЬ
        self.ui.MemoryTransportCost.setText(str(self.sumTransport))
        self.ui.MemoryVatSum_2.setText(str(self.totalInvoicePaymentOut))
        self.ui.tableWidget.clear()
        self.ui.tableWidget.setHorizontalHeaderLabels(('ИНВОЙС', "Вал.", "%", "Пошлина", "Вал.", "НДС", "Вал.", "Тр-т", "Вал." )) # т.к. занулили лог то переделываем наименование колонок

        self.data = []                          # нулим наш массив -лог
        i = 0                                   # и создаем заного
        for i in range(100):
            self.data.append([0]*10)

        self.data[0] = ['№','ИНВОЙС', "Вал.", "%", "Пошлина", "Вал.", "НДС", "Вал.", "Тр-т", "Вал."]


        return self.goodsNumber, self.invoiceSum, self.sumTax, self.sumVAT, self.data, self.uchet    # возвращаем нули обратно

    def pushButtonExcelExport(self):
        self.pushButtonMemory()


        wb = openpyxl.Workbook()    # вот так работаем через модуль с эксель

        # добавляем новый лист
        wb.create_sheet(title = 'Первый лист', index = 0)

        # получаем лист, с которым будем работать
        sheet = wb['Первый лист']
        #ws = wb.active
        ft = Font(italic=True, bold=True, size=12)
        ft2 = Font(italic=True, bold=True, size=20)
        row2 = 0
        row = 0
        col = 0

        for row in range(1, 100):
            if self.data[row-1][0] == 0:    #проверяем. если дошли сумма инвойса для записи в лог ноль значит дальше писать не надо 
                row=95                      
                col=1
            else:
                row2 = row                 #p запоминаем момент окончания данных 

            for col in range(1, 11):        # если нет - то пишем в эксель

                value = self.data[row-1][col-1]

                cell = sheet.cell(row = row, column = col)
                if col-1 == 1   and row-1 > 0:  
                    cell.value = float(value)       # причем, если хотим, чтобы в экселе были цифры - то преобразуем в цифры

                elif col-1 == 4   and row-1 > 0:  
                    cell.value = float(value)

                elif col-1 == 6   and row-1 > 0:  
                    cell.value = float(value)

                elif col-1 == 8   and row-1 > 0:  
                    cell.value = float(value)

                else:
                    cell.value = value




#        a1 = ws['A1']
#        d4 = ws['D4']

#        a1.font = ft
#        d4.font = ft

#        a1.font.italic = True # is not allowed # doctest: +SKIP
        # If you want to change the color of a Font, you need to reassign it::

#        a1.font = Font(color="FF0000", italic=True) # the change only affects A1



        cell = sheet.cell(row = row2+1, column = 2)
        value = "=SUM(B2:B"+str(row2)+")"
        cell.value = value        
        cell.font = ft

        cell = sheet.cell(row = row2+1, column = 3)
        value = "=C"+str(row2)
        cell.value = value        
        cell.font = ft


        cell = sheet.cell(row = row2+1, column = 5)
        value = "=SUM(E2:E"+str(row2)+")"
        cell.value = value        
        cell.font = ft

        cell = sheet.cell(row = row2+1, column = 6)
        value = "=F"+str(row2)
        cell.value = value        
        cell.font = ft



        cell = sheet.cell(row = row2+1, column = 7)
        value = "=SUM(G2:G"+str(row2)+")"
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2+1, column = 8)
        value = "=H"+str(row2)
        cell.value = value        
        cell.font = ft



        cell = sheet.cell(row = row2+1, column = 9)
        value = "=SUM(I2:I"+str(row2)+")"
        cell.value = value    
        cell.font = ft


        cell = sheet.cell(row = row2+1, column = 10)
        value = "=J"+str(row2)
        cell.value = value        
        cell.font = ft
    
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! пишем итог

        cell = sheet.cell(row = row2 + 3, column = 3)
        value = "Итого:"
        cell.value = value     
        cell.font = ft2

#Сборы пишем в итог
        cell = sheet.cell(row = row2 + 4, column = 3)
        value = "Сбор по инвойсу:"
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 4, column = 7)
        self.sborCurrency = round(self.sborCurrency,2)
        value = self.sborCurrency
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 4, column = 8)
        value = str(self.data[1][5])
        cell.value = value     
        cell.font = ft

#Пошлину пишем в итог
        cell = sheet.cell(row = row2 + 5, column = 3)
        value = "Пошлина по инвойсу:"
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 5, column = 7)
        self.sumTax = round(self.sumTax, 2)
        value = self.sumTax
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 5, column = 8)
        value = str(self.data[1][5])
        cell.value = value     
        cell.font = ft

        

#НДС пишем в итог

        cell = sheet.cell(row = row2 + 6, column = 3)
        value = "НДС по инвойсу:"
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 6, column = 7)
        self.sumVAT = round(self.sumVAT, 2)
        value = self.sumVAT
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 6, column = 8)
        value = str(self.data[1][7])
        cell.value = value     
        cell.font = ft



#Сумму платежей пишем в итог
        cell = sheet.cell(row = row2 + 8, column = 3)
        value = "Всего платежей по инвойсу:"
        cell.value = value     
        cell.font = ft


        cell = sheet.cell(row = row2 + 8, column = 7)
        self.totalInvoicePayment = round(self.totalInvoicePayment, 2)
        value = self.totalInvoicePayment
        cell.value = value     
        cell.font = ft

        cell = sheet.cell(row = row2 + 8, column = 8)
        value = str(self.data[1][7])
        cell.value = value     
        cell.font = ft



#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!



        cell = sheet.cell(row = row2+20, column = 1)     # когда закончили писать то через 2 строчки нишем текст ниже
        value = "Отчет создан программой ImportHelper (c)Slay " + str(self.now)
        cell.value = value

#        file_name = "report.xlsx"

        options = QFileDialog.Options()
        wb_patch = QtWidgets.QFileDialog.getSaveFileName(self,"QFileDialog.getSaveFileName()","","Excel files (*.xlsx);;All Files (*)", options=options)[0]
        file_name = wb_patch 
        try:
            wb.save(file_name)
            self.file_name = file_name.replace('/', chr(92))
            out = "Выгружен по пути: "+ self.file_name
            self.ui.filepathOut.setText(out) 
        except:
            out = "Ошибка записи"
            self.ui.filepathOut.setText(out)   

        return self.file_name
    
    def pushButtonOpenFile(self):
        os.startfile(self.file_name)


app = QtWidgets.QApplication([])
application = mywindow()
application.show()

app.setStyle('Fusion')  # для применения стилей цветов
sys.exit(app.exec())
